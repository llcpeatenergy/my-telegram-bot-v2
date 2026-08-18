import sys
import os
import openpyxl
from datetime import datetime, timedelta, timezone
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ConversationHandler, ContextTypes
import asyncio
import traceback

# ========================================
# ПРИМУСОВЕ ЛОГУВАННЯ
# ========================================
print("=" * 60)
print("🤖 БОТ ЗАПУСКАЄТЬСЯ")
print(f"Python версія: {sys.version}")
print(f"Поточна папка: {os.getcwd()}")
print("=" * 60)
sys.stdout.flush()

# ========================================
# НАЛАШТУВАННЯ
# ========================================
TOKEN = os.environ.get("TOKEN")
if not TOKEN:
    print("❌ ПОМИЛКА: Токен не знайдено в змінних середовища!")
    print("Додайте змінну TOKEN у налаштуваннях Koyeb")
    sys.exit(1)

EXCEL_FILE = "notebook.xlsx"
CHAT_ID_FILE = "chat_id.txt"
CHECK_INTERVAL = 60

# Київський час (UTC+3)
UKRAINE_TZ = timezone(timedelta(hours=3))

print(f"✅ Токен завантажено: {TOKEN[:10]}...")
print(f"✅ Часовий пояс: UTC+3 (Київ)")
sys.stdout.flush()

# ========================================
# ФУНКЦІЇ ДЛЯ РОБОТИ З CHAT_ID
# ========================================
def save_chat_id(chat_id):
    """Зберігає chat_id у файл"""
    try:
        with open(CHAT_ID_FILE, 'w') as f:
            f.write(str(chat_id))
        print(f"✅ chat_id {chat_id} збережено у файл")
        return True
    except Exception as e:
        print(f"❌ Помилка збереження chat_id: {e}")
        return False

def load_chat_id():
    """Завантажує chat_id з файлу"""
    try:
        if os.path.exists(CHAT_ID_FILE):
            with open(CHAT_ID_FILE, 'r') as f:
                chat_id = int(f.read().strip())
            print(f"✅ chat_id {chat_id} завантажено з файлу")
            return chat_id
        else:
            print("⚠️ Файл chat_id.txt не знайдено")
            return None
    except Exception as e:
        print(f"❌ Помилка завантаження chat_id: {e}")
        return None

# Завантажуємо chat_id при запуску
SAVED_CHAT_ID = load_chat_id()
if SAVED_CHAT_ID:
    print(f"✅ Бот запам'ятав ваш chat_id: {SAVED_CHAT_ID}")
else:
    print("⚠️ Бот ще не знає ваш chat_id. Напишіть /start")

sys.stdout.flush()

CATEGORIES = ["ТЕК", "Ідеї", "Особисті", "Книги", "Фільми", "Що відвідати", "Цікаві думки"]
TASK_CATEGORIES = ["ТЕК", "Ідеї", "Особисті"]
MEDIA_CATEGORIES = ["Книги", "Фільми", "Що відвідати"]
THOUGHT_CATEGORIES = ["Цікаві думки"]
CATEGORY, DESCRIPTION, PRIORITY, RESPONSIBLE, DUE_DATE, NAME, LINK = range(7)
REMINDER_TIME = 7
REMINDER_ACTION = 8

print("✅ Налаштування завантажено")
sys.stdout.flush()

# ========================================
# ДОПОМІЖНІ ФУНКЦІЇ ДЛЯ КНОПОК
# ========================================
def get_back_button(step):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Назад", callback_data=f"back_{step}")]
    ])

def get_skip_and_back_buttons(step):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏭️ Пропустити", callback_data=f"skip_{step}")],
        [InlineKeyboardButton("⬅️ Назад", callback_data=f"back_{step}")]
    ])

def get_category_keyboard():
    keyboard = [[InlineKeyboardButton(cat, callback_data=cat)] for cat in CATEGORIES]
    return InlineKeyboardMarkup(keyboard)

def get_reminder_action_keyboard(row_index):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏰ Перенагадати", callback_data=f"re_remind_{row_index}")],
        [InlineKeyboardButton("✅ Виконано", callback_data=f"done_remind_{row_index}")]
    ])

# ========================================
# РОБОТА З EXCEL
# ========================================
def get_headers():
    return ["Категорія", "Дата", "Опис", "Пріоритет", "Хто відповідальний", 
            "Дата виконання", "Назва", "Лінк", "Нагадати о"]

def get_now_ukraine():
    return datetime.now(UKRAINE_TZ)

def save_to_excel(data):
    try:
        print(f"📝 Зберігаю в Excel: {data.get('description', '')[:30]}...")
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sheet = wb.active
        except FileNotFoundError:
            print("📄 Створюю новий файл Excel...")
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(get_headers())
        
        now_ukraine = get_now_ukraine()
        
        row = [
            data.get("category", ""),
            now_ukraine.strftime("%Y-%m-%d %H:%M:%S"),
            data.get("description", ""),
            data.get("priority", ""),
            data.get("responsible", ""),
            data.get("due_date", ""),
            data.get("name", ""),
            data.get("link", ""),
            data.get("reminder_time", "")
        ]
        sheet.append(row)
        wb.save(EXCEL_FILE)
        print("✅ Excel збережено!")
        sys.stdout.flush()
        return True
    except Exception as e:
        print(f"❌ Помилка збереження Excel: {e}")
        sys.stdout.flush()
        return False

def update_reminder_time(row_index, new_time):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        reminder_col = None
        for i, h in enumerate(headers):
            if h == "Нагадати о":
                reminder_col = i + 1
                break
        if reminder_col:
            row = sheet[row_index + 2]
            row[reminder_col - 1].value = new_time
            wb.save(EXCEL_FILE)
            print(f"✅ Час нагадування оновлено: {new_time}")
            sys.stdout.flush()
            return True
    except Exception as e:
        print(f"❌ Помилка оновлення нагадування: {e}")
        sys.stdout.flush()
    return False

def clear_reminder(row_index):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        reminder_col = None
        for i, h in enumerate(headers):
            if h == "Нагадати о":
                reminder_col = i + 1
                break
        if reminder_col:
            row = sheet[row_index + 2]
            row[reminder_col - 1].value = None
            wb.save(EXCEL_FILE)
            print("✅ Нагадування очищено")
            sys.stdout.flush()
            return True
    except Exception as e:
        print(f"❌ Помилка очищення нагадування: {e}")
        sys.stdout.flush()
    return False

def check_reminders():
    reminders = []
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        reminder_col = None
        for i, h in enumerate(headers):
            if h == "Нагадати о":
                reminder_col = i + 1
                break
        if not reminder_col:
            return reminders
        
        now_ukraine = get_now_ukraine()
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False)):
            reminder_value = row[reminder_col - 1].value
            if reminder_value and isinstance(reminder_value, str):
                try:
                    reminder_time = datetime.strptime(reminder_value, "%Y-%m-%d %H:%M")
                    reminder_time = reminder_time.replace(tzinfo=UKRAINE_TZ)
                    if reminder_time <= now_ukraine:
                        reminders.append({
                            "row_index": idx,
                            "row": row,
                            "category": row[0].value or "",
                            "description": row[2].value or "",
                            "priority": row[3].value or "",
                            "responsible": row[4].value or "",
                            "reminder_time": reminder_value
                        })
                except:
                    pass
    except Exception as e:
        print(f"❌ Помилка перевірки нагадувань: {e}")
        sys.stdout.flush()
    return reminders

# ========================================
# ФОНОВИЙ ПРОЦЕС ДЛЯ НАГАДУВАНЬ
# ========================================
async def reminder_loop(app: Application):
    while True:
        try:
            reminders = check_reminders()
            if reminders:
                for reminder in reminders:
                    try:
                        chat_id = load_chat_id()
                        if not chat_id:
                            chat_id = app.bot_data.get("chat_id")
                        
                        if chat_id:
                            message = (
                                f"⏰ **Нагадування!**\n\n"
                                f"📌 Категорія: {reminder['category']}\n"
                                f"📝 Опис: {reminder['description']}\n"
                                f"⚡ Пріоритет: {reminder['priority']}\n"
                                f"👤 Відповідальний: {reminder.get('responsible', '—')}\n"
                                f"⏳ Час: {reminder['reminder_time']}"
                            )
                            await app.bot.send_message(
                                chat_id=chat_id,
                                text=message,
                                reply_markup=get_reminder_action_keyboard(reminder['row_index']),
                                parse_mode="Markdown"
                            )
                            print("✅ Нагадування надіслано в чат")
                            clear_reminder(reminder['row_index'])
                        else:
                            print("⚠️ chat_id не знайдено. Нагадування не надіслано.")
                    except Exception as e:
                        print(f"❌ Помилка надсилання нагадування: {e}")
        except Exception as e:
            print(f"❌ Помилка в reminder_loop: {e}")
        await asyncio.sleep(CHECK_INTERVAL)

# ========================================
# КОМАНДА ЗАВАНТАЖЕННЯ EXCEL
# ========================================
async def download_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if os.path.exists(EXCEL_FILE):
            await update.message.reply_document(
                document=open(EXCEL_FILE, 'rb'),
                filename="notebook.xlsx",
                caption="📊 Ось ваш файл з нотатками!"
            )
        else:
            await update.message.reply_text("❌ Файл ще не створено. Зробіть хоча б один запис.")
    except Exception as e:
        await update.message.reply_text(f"❌ Помилка: {e}")

# ========================================
# ОБРОБНИК ПЕРЕНАГАДУВАННЯ
# ========================================
async def handle_reminder_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    data = query.data
    if data.startswith("done_remind_"):
        row_index = int(data.split("_")[2])
        if clear_reminder(row_index):
            await query.edit_message_text("✅ Нагадування виконано! Запис очищено.")
        else:
            await query.edit_message_text("❌ Помилка при очищенні нагадування.")
        return ConversationHandler.END
    
    elif data.startswith("re_remind_"):
        row_index = int(data.split("_")[2])
        context.user_data["remind_row_index"] = row_index
        await query.edit_message_text(
            "⏰ Введіть нову дату та час нагадування у форматі:\n"
            "`ДД.ММ.РРРР ГГ:ХХ`\n\n"
            "Наприклад: `20.08.2026 15:00`",
            parse_mode="Markdown"
        )
        return REMINDER_ACTION

async def handle_new_reminder_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    row_index = context.user_data.get("remind_row_index")
    if row_index is None:
        await update.message.reply_text("❌ Помилка: не знайдено рядок для оновлення.")
        return ConversationHandler.END
    
    reminder_text = update.message.text.strip()
    try:
        reminder_time = datetime.strptime(reminder_text, "%d.%m.%Y %H:%M")
        new_time_str = reminder_time.strftime("%Y-%m-%d %H:%M")
        if update_reminder_time(row_index, new_time_str):
            await update.message.reply_text(
                f"✅ Нагадування перенесено на {reminder_text}"
            )
        else:
            await update.message.reply_text("❌ Помилка при оновленні нагадування.")
        context.user_data.pop("remind_row_index", None)
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text(
            "❌ Неправильний формат. Введіть дату та час у форматі:\n"
            "`ДД.ММ.РРРР ГГ:ХХ`\n\n"
            "Наприклад: `20.08.2026 15:00`",
            parse_mode="Markdown"
        )
        return REMINDER_ACTION

# ========================================
# ОСНОВНІ ФУНКЦІЇ
# ========================================
async def start(update, context):
    chat_id = update.effective_chat.id
    save_chat_id(chat_id)
    context.bot_data["chat_id"] = chat_id
    
    print(f"📩 /start від {update.effective_user.username}, chat_id: {chat_id}")
    sys.stdout.flush()
    
    await update.message.reply_text(
        "📋 Оберіть категорію:",
        reply_markup=get_category_keyboard()
    )
    return CATEGORY

async def handle_text(update, context):
    if "category" not in context.user_data:
        await update.message.reply_text(
            "⚠️ Спочатку оберіть категорію за допомогою кнопок.",
            reply_markup=get_category_keyboard()
        )
        return CATEGORY
    return await start(update, context)

async def category_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    category = query.data
    context.user_data["category"] = category
    print(f"📂 Вибрано: {category}")
    sys.stdout.flush()
    
    if category in TASK_CATEGORIES:
        await query.edit_message_text(
            f"📌 {category}\n\n✏️ Введіть ОПИС:",
            reply_markup=get_back_button("category")
        )
        return DESCRIPTION
        
    elif category in MEDIA_CATEGORIES:
        await query.edit_message_text(
            f"📌 {category}\n\n📝 Введіть НАЗВУ:",
            reply_markup=get_back_button("category")
        )
        return NAME
        
    elif category in THOUGHT_CATEGORIES:
        await query.edit_message_text(
            f"📌 {category}\n\n🧠 Напишіть думку:",
            reply_markup=get_back_button("category")
        )
        return DESCRIPTION
        
    await query.edit_message_text("❌ Помилка.")
    return ConversationHandler.END

async def get_description(update, context):
    if "category" not in context.user_data:
        await update.message.reply_text(
            "⚠️ Спочатку оберіть категорію за допомогою кнопок.",
            reply_markup=get_category_keyboard()
        )
        return CATEGORY
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        if query.data == "back_category":
            await query.edit_message_text(
                "📋 Оберіть категорію:",
                reply_markup=get_category_keyboard()
            )
            return CATEGORY
        return DESCRIPTION
    
    context.user_data["description"] = update.message.text
    category = context.user_data["category"]
    
    if category in THOUGHT_CATEGORIES:
        save_to_excel(context.user_data)
        await update.message.reply_text("✅ Збережено!")
        context.user_data.clear()
        return ConversationHandler.END
    
    keyboard = [
        [InlineKeyboardButton("🔴 Високий", callback_data="Високий")],
        [InlineKeyboardButton("🟡 Середній", callback_data="Середній")],
        [InlineKeyboardButton("🟢 Низький", callback_data="Низький")],
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_priority")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("⚡ Пріоритет:", reply_markup=reply_markup)
    return PRIORITY

async def get_priority(update, context):
    query = update.callback_query
    await query.answer()
    
    if query.data == "back_priority":
        await query.edit_message_text(
            f"📌 {context.user_data.get('category', '')}\n\n✏️ Введіть ОПИС:",
            reply_markup=get_back_button("category")
        )
        return DESCRIPTION
    
    context.user_data["priority"] = query.data
    await query.edit_message_text(
        "👤 Відповідальний:",
        reply_markup=get_skip_and_back_buttons("responsible")
    )
    return RESPONSIBLE

async def get_responsible(update, context):
    if "category" not in context.user_data:
        await update.message.reply_text(
            "⚠️ Оберіть категорію:",
            reply_markup=get_category_keyboard()
        )
        return CATEGORY
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        
        if query.data == "back_responsible":
            keyboard = [
                [InlineKeyboardButton("🔴 Високий", callback_data="Високий")],
                [InlineKeyboardButton("🟡 Середній", callback_data="Середній")],
                [InlineKeyboardButton("🟢 Низький", callback_data="Низький")],
                [InlineKeyboardButton("⬅️ Назад", callback_data="back_priority")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text("⚡ Пріоритет:", reply_markup=reply_markup)
            return PRIORITY
        
        elif query.data == "skip_responsible":
            context.user_data["responsible"] = ""
            await query.edit_message_text(
                "📅 Дата виконання:",
                reply_markup=get_skip_and_back_buttons("due_date")
            )
            return DUE_DATE
    
    context.user_data["responsible"] = update.message.text
    await update.message.reply_text(
        "📅 Дата виконання:",
        reply_markup=get_skip_and_back_buttons("due_date")
    )
    return DUE_DATE

async def get_due_date(update, context):
    if "category" not in context.user_data:
        await update.message.reply_text(
            "⚠️ Оберіть категорію:",
            reply_markup=get_category_keyboard()
        )
        return CATEGORY
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        
        if query.data == "back_due_date":
            await query.edit_message_text(
                "👤 Відповідальний:",
                reply_markup=get_skip_and_back_buttons("responsible")
            )
            return RESPONSIBLE
        
        elif query.data == "skip_due_date":
            context.user_data["due_date"] = ""
            category = context.user_data.get("category", "")
            if category in TASK_CATEGORIES:
                await query.edit_message_text(
                    "⏰ Введіть дату та час нагадування (наприклад, 20.08.2026 15:00)\nабо натисніть 'Пропустити':",
                    reply_markup=get_skip_and_back_buttons("reminder")
                )
                return REMINDER_TIME
            else:
                save_to_excel(context.user_data)
                await query.edit_message_text("✅ Збережено!")
                context.user_data.clear()
                return ConversationHandler.END
    
    context.user_data["due_date"] = update.message.text
    category = context.user_data.get("category", "")
    if category in TASK_CATEGORIES:
        await update.message.reply_text(
            "⏰ Введіть дату та час нагадування (наприклад, 20.08.2026 15:00)\nабо натисніть 'Пропустити':",
            reply_markup=get_skip_and_back_buttons("reminder")
        )
        return REMINDER_TIME
    else:
        save_to_excel(context.user_data)
        await update.message.reply_text("✅ Збережено!")
        context.user_data.clear()
        return ConversationHandler.END

async def get_reminder_time(update, context):
    if "category" not in context.user_data:
        await update.message.reply_text(
            "⚠️ Оберіть категорію:",
            reply_markup=get_category_keyboard()
        )
        return CATEGORY
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        
        if query.data == "back_reminder":
            await query.edit_message_text(
                "📅 Дата виконання:",
                reply_markup=get_skip_and_back_buttons("due_date")
            )
            return DUE_DATE
        
        elif query.data == "skip_reminder":
            context.user_data["reminder_time"] = ""
            save_to_excel(context.user_data)
            await query.edit_message_text("✅ Збережено!")
            context.user_data.clear()
            return ConversationHandler.END
    
    reminder_text = update.message.text.strip()
    try:
        local_time = datetime.strptime(reminder_text, "%d.%m.%Y %H:%M")
        context.user_data["reminder_time"] = local_time.strftime("%Y-%m-%d %H:%M")
        save_to_excel(context.user_data)
        await update.message.reply_text(
            f"✅ Збережено! Нагадування встановлено на {reminder_text} (за вашим часом)"
        )
        context.user_data.clear()
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text(
            "❌ Неправильний формат. Введіть дату та час у форматі ДД.ММ.РРРР ГГ:ХХ\nНаприклад: 20.08.2026 15:00"
        )
        return REMINDER_TIME

async def get_name(update, context):
    if "category" not in context.user_data:
        await update.message.reply_text(
            "⚠️ Оберіть категорію:",
            reply_markup=get_category_keyboard()
        )
        return CATEGORY
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        if query.data == "back_category":
            await query.edit_message_text(
                "📋 Оберіть категорію:",
                reply_markup=get_category_keyboard()
            )
            return CATEGORY
        return NAME
    
    context.user_data["name"] = update.message.text
    await update.message.reply_text(
        "🔗 Лінк:",
        reply_markup=get_skip_and_back_buttons("link")
    )
    return LINK

async def get_link(update, context):
    if "category" not in context.user_data:
        await update.message.reply_text(
            "⚠️ Оберіть категорію:",
            reply_markup=get_category_keyboard()
        )
        return CATEGORY
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        
        if query.data == "back_link":
            await query.edit_message_text(
                f"📌 {context.user_data.get('category', '')}\n\n📝 Введіть НАЗВУ:",
                reply_markup=get_back_button("category")
            )
            return NAME
        
        elif query.data == "skip_link":
            context.user_data["link"] = ""
            save_to_excel(context.user_data)
            await query.edit_message_text("✅ Збережено!")
            context.user_data.clear()
            return ConversationHandler.END
    
    context.user_data["link"] = update.message.text
    save_to_excel(context.user_data)
    await update.message.reply_text("✅ Збережено!")
    context.user_data.clear()
    return ConversationHandler.END

async def cancel(update, context):
    context.user_data.clear()
    await update.message.reply_text("❌ Скасовано.")
    return ConversationHandler.END

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print(f"❌ Помилка: {context.error}")
    sys.stdout.flush()
    if isinstance(context.error, Exception):
        error_str = str(context.error)
        if "NetworkError" in error_str or "Conflict" in error_str:
            print("⚠️ Проблема з мережею. Бот продовжує роботу...")
            sys.stdout.flush()
            return
    if update and update.effective_message:
        await update.effective_message.reply_text(
            "❌ Сталася помилка. Спробуйте ще раз через /start"
        )

# ========================================
# ЗАПУСК
# ========================================
def main():
    try:
        print("📦 Створення додатку...")
        sys.stdout.flush()
        app = Application.builder().token(TOKEN).build()
        print("✅ Додаток створено")
        sys.stdout.flush()
        
        app.add_error_handler(error_handler)
        app.add_handler(CommandHandler("download", download_excel))
        print("✅ Команду /download додано")
        sys.stdout.flush()
        
        app.add_handler(CallbackQueryHandler(handle_reminder_action, pattern="^(done_remind_|re_remind_)"))
        print("✅ Обробник перенагадування додано")
        sys.stdout.flush()
        
        conv_handler = ConversationHandler(
            entry_points=[
                CommandHandler("start", start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text)
            ],
            states={
                CATEGORY: [CallbackQueryHandler(category_selected)],
                DESCRIPTION: [
                    CallbackQueryHandler(get_description),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, get_description)
                ],
                PRIORITY: [CallbackQueryHandler(get_priority)],
                RESPONSIBLE: [
                    CallbackQueryHandler(get_responsible),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, get_responsible)
                ],
                DUE_DATE: [
                    CallbackQueryHandler(get_due_date),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, get_due_date)
                ],
                REMINDER_TIME: [
                    CallbackQueryHandler(get_reminder_time),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, get_reminder_time)
                ],
                REMINDER_ACTION: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, handle_new_reminder_time)
                ],
                NAME: [
                    CallbackQueryHandler(get_name),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)
                ],
                LINK: [
                    CallbackQueryHandler(get_link),
                    MessageHandler(filters.TEXT & ~filters.COMMAND, get_link)
                ],
            },
            fallbacks=[CommandHandler("cancel", cancel)],
        )
        app.add_handler(conv_handler)
        print("✅ ConversationHandler додано")
        sys.stdout.flush()
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.create_task(reminder_loop(app))
        print("✅ Фоновий процес нагадувань запущено")
        sys.stdout.flush()
        
        print("=" * 60)
        print("🤖 БОТ УСПІШНО ЗАПУЩЕНО!")
        print("📌 Команди:")
        print("  /start - почати новий запис")
        print("  /download - завантажити Excel-файл")
        print("⏰ Нагадування перевіряються кожні 60 секунд")
        print("=" * 60)
        sys.stdout.flush()
        
        app.run_polling(allowed_updates=Update.ALL_TYPES)
        
    except Exception as e:
        print(f"❌ КРИТИЧНА ПОМИЛКА: {e}")
        traceback.print_exc()
        sys.stdout.flush()

if __name__ == "__main__":
    main()
    print("⚠️ Бот завершив роботу")
    sys.stdout.flush()
